from html import escape
from pathlib import Path

from openpyxl import load_workbook


xlsx = Path("downloads/friday_referral_20260501_20260521_generated.xlsx")
out_html = Path("downloads/friday_referral_20260501_20260521_clipboard.html")
out_txt = Path("downloads/friday_referral_20260501_20260521_clipboard.txt")

wb = load_workbook(xlsx, data_only=True)

notes = """
<h3>2、本周的达成情况（5.1-5.21）</h3>
<p><strong>分团队达成情况</strong></p>
<p>💡 欧美澳问题：CC01组约课率60.00%，CC11组约课率50.00%；CC16组到课转化率16.67%，例子转化率4.08%；5.19-25批次多个小组到课转化和例子转化偏低。</p>
<p>港澳问题：港澳CC02组约课率36.84%，到课转化率0.00%，例子转化率0.00%；港澳CC03组约课率63.64%，到课转化率0.00%，例子转化率0.00%。</p>
"""

css = """
<style>
table{border-collapse:collapse;font-family:Arial,"Microsoft YaHei",sans-serif;font-size:12px;margin:8px 0 18px 0;}
td,th{border:1px solid #000;padding:4px 6px;text-align:center;white-space:nowrap;}
th{font-weight:bold;background:#8EAADB;}
th.metric{background:#F8CBAD;}
.low{background:#F4B6B6;}
h4{margin:14px 0 6px 0;}
</style>
"""


def fmt(value, col):
    if value is None:
        return ""
    if col >= 9 and isinstance(value, (int, float)):
        return f"{value:.2%}"
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value)


def sheet_html(ws):
    parts = [f"<h4>{escape(ws.title)}</h4><table>"]
    for row_num, row in enumerate(ws.iter_rows(values_only=True), start=1):
        if row_num == 1:
            continue
        tag = "th" if row_num == 2 else "td"
        parts.append("<tr>")
        for col_num, value in enumerate(row, start=1):
            cls = ""
            if row_num == 2 and col_num >= 9:
                cls = ' class="metric"'
            elif row_num > 2 and col_num >= 9 and isinstance(value, (int, float)):
                if (col_num in (12, 13) and value < 0.25) or (
                    col_num in (9, 10, 11) and value < 0.70
                ):
                    cls = ' class="low"'
            parts.append(f"<{tag}{cls}>{escape(fmt(value, col_num))}</{tag}>")
        parts.append("</tr>")
    parts.append("</table>")
    return "\n".join(parts)


html = (
    "<html><head>"
    + css
    + "</head><body>"
    + notes
    + "".join(sheet_html(wb[sheet_name]) for sheet_name in wb.sheetnames)
    + "</body></html>"
)
out_html.write_text(html, encoding="utf-8")

plain = [
    "2、本周的达成情况（5.1-5.21）",
    "分团队达成情况",
    "欧美澳问题：CC01组约课率60.00%，CC11组约课率50.00%；CC16组到课转化率16.67%，例子转化率4.08%。",
    "港澳问题：港澳CC02组约课率36.84%，港澳CC03组约课率63.64%，到课转化率和例子转化率偏低。",
]
out_txt.write_text("\n".join(plain), encoding="utf-8")

print(out_html)
print(out_txt)

