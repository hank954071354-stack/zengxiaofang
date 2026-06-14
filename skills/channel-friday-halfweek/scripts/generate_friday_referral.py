from pathlib import Path
from collections import OrderedDict

import openpyxl
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side


ROOT = Path.cwd()
DOWNLOADS = ROOT / "downloads"
OUT_XLSX = DOWNLOADS / "friday_referral_20260501_20260521_generated.xlsx"
OUT_MD = DOWNLOADS / "friday_referral_20260501_20260521_notes.md"


def find_source() -> Path:
    candidates = sorted(
        DOWNLOADS.glob("*20260522*/*.xlsx"),
        key=lambda p: p.stat().st_mtime,
        reverse=True,
    )
    if not candidates:
        raise FileNotFoundError("No 20260522 SmartBI xlsx found under downloads.")
    return candidates[0]


def rate(num, den):
    return None if not den else num / den


def row_from_totals(batch, typ, group, totals):
    first, appt, absent, attend, deal = totals
    return [
        batch,
        typ,
        group,
        int(first),
        int(appt),
        int(absent),
        int(attend),
        int(deal),
        rate(appt, first),
        rate(attend, appt),
        rate(attend, appt),
        rate(deal, attend),
        rate(deal, first),
    ]


def load_records(src: Path):
    wb = load_workbook(src, data_only=True)
    ws = wb["手推&非手推"]
    records = []
    current_batch = ""
    current_type = ""

    for raw in ws.iter_rows(min_row=5, values_only=True):
        vals = list(raw[1:15])
        if not any(v is not None and v != "" for v in vals):
            continue

        batch, typ, group = vals[0], vals[1], vals[2]
        if batch:
            current_batch = str(batch)
        if typ:
            current_type = str(typ)

        group = str(group or "")
        if not group or group == "总计":
            continue

        nums = []
        for value in vals[4:9]:
            try:
                nums.append(float(value or 0))
            except Exception:
                nums.append(0.0)

        records.append(
            {
                "批次": current_batch,
                "类型": current_type,
                "首发CC小组": group,
                "首发学员数": nums[0],
                "最近一次体验课约课学员数": nums[1],
                "缺席学员数": nums[2],
                "最近一次体验课到课学员数": nums[3],
                "成交学员数": nums[4],
            }
        )
    return records


def build_region(records, predicate):
    data = [row for row in records if predicate(row["首发CC小组"])]
    batches = list(OrderedDict((row["批次"], None) for row in data).keys())
    rows = []

    for batch in batches:
        for typ in ["手推", "非手推"]:
            groups = sorted(
                {
                    row["首发CC小组"]
                    for row in data
                    if row["批次"] == batch and row["类型"] == typ
                }
            )
            if not groups:
                continue

            totals = [0, 0, 0, 0, 0]
            group_rows = []
            for group in groups:
                group_totals = [0, 0, 0, 0, 0]
                for row in data:
                    if (
                        row["批次"] == batch
                        and row["类型"] == typ
                        and row["首发CC小组"] == group
                    ):
                        values = [
                            row["首发学员数"],
                            row["最近一次体验课约课学员数"],
                            row["缺席学员数"],
                            row["最近一次体验课到课学员数"],
                            row["成交学员数"],
                        ]
                        group_totals = [
                            group_totals[index] + values[index] for index in range(5)
                        ]
                totals = [totals[index] + group_totals[index] for index in range(5)]
                group_rows.append(row_from_totals("", "", group, group_totals))

            rows.append(row_from_totals(batch, typ, "总计", totals))
            rows.extend(group_rows)

    for typ in ["手推", "非手推"]:
        groups = sorted({row["首发CC小组"] for row in data if row["类型"] == typ})
        if not groups:
            continue

        totals = [0, 0, 0, 0, 0]
        group_rows = []
        for group in groups:
            group_totals = [0, 0, 0, 0, 0]
            for row in data:
                if row["类型"] == typ and row["首发CC小组"] == group:
                    values = [
                        row["首发学员数"],
                        row["最近一次体验课约课学员数"],
                        row["缺席学员数"],
                        row["最近一次体验课到课学员数"],
                        row["成交学员数"],
                    ]
                    group_totals = [
                        group_totals[index] + values[index] for index in range(5)
                    ]
            totals = [totals[index] + group_totals[index] for index in range(5)]
            group_rows.append(row_from_totals("", "", group, group_totals))

        rows.append(row_from_totals("总计", typ, "总计", totals))
        rows.extend(group_rows)

    return rows


def write_workbook(regions):
    wb = Workbook()
    wb.remove(wb.active)

    header = [
        "批次",
        "类型",
        "首发CC小组",
        "首发学员数",
        "最近一次体验课约课学员数",
        "缺席学员数",
        "最近一次体验课到课学员数",
        "成交学员数",
        "约课率",
        "约课到课率",
        "应到课率",
        "到课转化率",
        "例子转化率",
    ]
    blue = PatternFill("solid", fgColor="8EAADB")
    orange = PatternFill("solid", fgColor="F8CBAD")
    red = PatternFill("solid", fgColor="F4B6B6")
    thin = Side(style="thin", color="000000")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for name, rows in regions.items():
        ws = wb.create_sheet(name[:31])
        ws.append([name])
        ws.append(header)

        for cell in ws[2]:
            cell.font = Font(bold=True)
            cell.fill = orange if cell.column >= 9 else blue
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = border

        for row in rows:
            ws.append(row)

        for row in ws.iter_rows(min_row=3, max_row=ws.max_row, max_col=len(header)):
            for cell in row:
                cell.border = border
                cell.alignment = Alignment(horizontal="center", vertical="center")
            for cell in row[8:]:
                if isinstance(cell.value, float):
                    cell.number_format = "0.00%"
                    if cell.column in (12, 13) and cell.value < 0.25:
                        cell.fill = red
                    if cell.column in (9, 10, 11) and cell.value < 0.70:
                        cell.fill = red

        widths = [18, 10, 18, 12, 24, 12, 24, 12, 12, 14, 12, 14, 14]
        for index, width in enumerate(widths, 1):
            ws.column_dimensions[openpyxl.utils.get_column_letter(index)].width = width
        ws.freeze_panes = "A3"

    wb.save(OUT_XLSX)


def pct(value):
    return "" if value is None else f"{value * 100:.2f}%"


def write_notes(regions):
    lines = [
        "# 2、本周的达成情况（5.1-5.21）",
        "",
        "分团队达成情况",
        "",
    ]
    for name, rows in regions.items():
        latest = next(
            (
                row[0]
                for row in rows
                if isinstance(row[0], str) and row[0].startswith("2026-05-19")
            ),
            "",
        )
        issues = []
        in_latest = False
        for row in rows:
            batch, typ, group = row[0], row[1], row[2]
            if batch == latest:
                in_latest = True
            elif batch and batch != latest:
                in_latest = False
            if not in_latest or group == "总计":
                continue

            prefix = f"{typ}{group}" if typ else group
            if row[8] is not None and row[8] < 0.70:
                issues.append(f"{prefix}约课率{pct(row[8])}")
            if row[11] is not None and row[11] < 0.25:
                issues.append(f"{prefix}到课转化率{pct(row[11])}")
            if row[12] is not None and row[12] < 0.10:
                issues.append(f"{prefix}例子转化率{pct(row[12])}")

        lines.append(f"## {name}")
        if issues:
            lines.append("问题：" + "；".join(issues[:10]) + "。")
        else:
            lines.append("问题：本周暂无明显低值异常，建议结合目标线复核。")
        lines.append("")
        lines.append(f"表格来源：{OUT_XLSX.name} / sheet：{name[:31]}")
        lines.append("")

    OUT_MD.write_text("\n".join(lines), encoding="utf-8")


def main():
    src = find_source()
    records = load_records(src)
    regions = OrderedDict(
        [
            (
                "欧美澳（剔除港澳台）",
                build_region(records, lambda group: "港澳" not in group and "台湾" not in group),
            ),
            ("港澳", build_region(records, lambda group: "港澳" in group)),
        ]
    )
    write_workbook(regions)
    write_notes(regions)
    print(OUT_XLSX)
    print(OUT_MD)


if __name__ == "__main__":
    main()

