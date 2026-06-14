from pathlib import Path
import json
import shutil
import tempfile

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side


BASE_DIR = Path(__file__).resolve().parent
SOURCE = sorted(
    [p for p in BASE_DIR.glob("*.xlsx") if "销售明细" in p.name and p.stat().st_size > 500_000],
    key=lambda p: p.stat().st_mtime,
    reverse=True,
)[0]
TEMPLATE = BASE_DIR / "【营销加码赠课】成本归销售-益智海外加码赠课活动-模板.xls"
TARGET_XLS = BASE_DIR / "【营销加码赠课】成本归销售-益智海外加码赠课活动-20260521.xls"
TARGET_XLSX = BASE_DIR / "【营销加码赠课】成本归销售-益智海外加码赠课活动-20260521.xlsx"
AUDIT = BASE_DIR / "加码赠课计算明细_20260521.xlsx"

# BI export columns, zero-based after row 8 header.
USECOLS = [1, 16, 27, 73, 111, 112, 113, 116, 128, 129]
NAMES = [
    "student_id",
    "area_level",
    "channel_cat",
    "deal_group",
    "order_no",
    "sign_time",
    "sign_amount",
    "is_installment",
    "refund_time",
    "refund_amount",
]


def clean_text(value):
    if pd.isna(value):
        return ""
    text = str(value).strip()
    return text[:-2] if text.endswith(".0") else text


def build_issue_workbook_from_template(rows):
    headers = [
        "豌豆/魔力用户id*",
        "学科品类*",
        "发放虚拟币数量*",
        "发放原因*",
        "发放备注*",
        "部门归属：第一级id*",
        "部门归属：最后一级id*",
        "部门归属：发放类型*",
        "关联的订单号",
        "是否合同内",
    ]
    temp_xlsx = Path(tempfile.gettempdir()) / "marketing_bonus_template_20260521.xlsx"
    shutil.copyfile(TEMPLATE, temp_xlsx)
    wb = load_workbook(temp_xlsx)
    ws = wb.active

    for col_idx, header in enumerate(headers, start=1):
        ws.cell(row=1, column=col_idx).value = header

    if ws.max_row > 1:
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, max_col=ws.max_column):
            for cell in row:
                cell.value = None

    template_row = 2 if ws.max_row >= 2 else 1
    for row_idx, row in enumerate(rows, start=2):
        values = [
            row["student_id"],
            "VIP_WanDou",
            row["coins"],
            "市场活动赠送",
            "海外益智销售营销活动发放-",
            5672,
            6360,
            16,
            row["order_no"],
            "是",
        ]
        for col_idx, value in enumerate(values, start=1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.value = value
            if row_idx > ws.max_row:
                src = ws.cell(row=template_row, column=col_idx)
                cell._style = src._style
                cell.number_format = src.number_format
                cell.alignment = src.alignment

    if ws.max_row > len(rows) + 1:
        ws.delete_rows(len(rows) + 2, ws.max_row - len(rows) - 1)
    return wb


def main():
    df = pd.read_excel(SOURCE, header=7, usecols=USECOLS, engine="openpyxl", dtype=object)
    df.columns = NAMES
    for col in ["student_id", "area_level", "channel_cat", "deal_group", "order_no", "is_installment"]:
        df[col] = df[col].astype("string").str.strip()

    df["student_id_clean"] = df["student_id"].map(clean_text)
    df["order_no_clean"] = df["order_no"].map(clean_text)
    df["sign_dt"] = pd.to_datetime(df["sign_time"], errors="coerce")
    df["amount"] = pd.to_numeric(df["sign_amount"], errors="coerce")
    df["refund_amt"] = pd.to_numeric(df["refund_amount"], errors="coerce").fillna(0)
    refund_time = df["refund_time"].astype("string").str.strip()
    not_refund = (
        refund_time.isna()
        | (refund_time == "")
        | (refund_time == "<NA>")
        | (refund_time.str.lower() == "nan")
    ) & (df["refund_amt"] <= 0)

    tw_groups = {"台湾CC01组", "台湾CC02组"}
    tw_period = df["sign_dt"].between(pd.Timestamp("2026-05-13"), pd.Timestamp("2026-05-19 23:59:59"))
    tw_base = not_refund & tw_period & df["deal_group"].isin(tw_groups)
    installment = df["is_installment"].fillna("").str.replace(r"\.0$", "", regex=True)
    non_installment = installment.isin(["", "0", "否", "False", "false", "nan", "<NA>"])
    amount_map = {3850: 0, 5280: 32000, 8880: 32000, 14550: 40000}
    amount_round = df["amount"].round().astype("Int64")
    tw1 = tw_base & non_installment & amount_round.isin(list(amount_map.keys()))
    tw2 = tw_base & (df["channel_cat"] == "投放")

    intl_groups = {"CC01组", "CC02组", "CC11组", "CC16组", "港澳CC01组", "港澳CC02组", "港澳CC03组"}
    intl_period = df["sign_dt"].between(pd.Timestamp("2026-05-15"), pd.Timestamp("2026-05-18 23:59:59"))
    business_channels = {
        "海外商务",
        "港澳商务",
        "Local商务",
        "LOCAL商务",
        "台湾商务",
        "海外商务",
        "海外港澳商务",
        "海外LOCAL商务",
        "海外Local商务",
        "海外台湾商务",
    }
    intl = (
        not_refund
        & intl_period
        & (df["area_level"] != "台湾")
        & df["deal_group"].isin(intl_groups)
        & ~df["channel_cat"].fillna("").isin(business_channels)
    )

    entries = []
    audit_rows = []

    def add(row, source_name, coins):
        item = {
            "source": source_name,
            "student_id": clean_text(row["student_id_clean"]),
            "coins": int(coins),
            "order_no": clean_text(row["order_no_clean"]),
        }
        entries.append(item)
        audit_rows.append(
            {
                **item,
                "sign_time": row["sign_time"],
                "sign_amount": row["amount"],
                "area_level": row["area_level"],
                "deal_group": row["deal_group"],
                "channel_cat": row["channel_cat"],
                "refund_time": row["refund_time"],
                "refund_amount": row["refund_amount"],
            }
        )

    for _, row in df[tw1].iterrows():
        amount = int(round(float(row["amount"]))) if pd.notna(row["amount"]) else 0
        coins = amount_map.get(amount, 0)
        if coins > 0:
            add(row, "台湾方案1", coins)
        else:
            audit_rows.append(
                {
                    "source": "台湾方案1-0币",
                    "student_id": clean_text(row["student_id_clean"]),
                    "coins": 0,
                    "order_no": clean_text(row["order_no_clean"]),
                    "sign_time": row["sign_time"],
                    "sign_amount": row["amount"],
                    "area_level": row["area_level"],
                    "deal_group": row["deal_group"],
                    "channel_cat": row["channel_cat"],
                    "refund_time": row["refund_time"],
                    "refund_amount": row["refund_amount"],
                }
            )
    for _, row in df[tw2].iterrows():
        add(row, "台湾方案2", 32000)
    for _, row in df[intl].iterrows():
        add(row, "欧美澳港澳", 16000)

    shutil.copyfile(TEMPLATE, TARGET_XLS)
    wb = build_issue_workbook_from_template(entries)
    wb.save(TARGET_XLSX)
    shutil.copyfile(TARGET_XLSX, TARGET_XLS)

    with pd.ExcelWriter(AUDIT, engine="openpyxl") as writer:
        pd.DataFrame(audit_rows).to_excel(writer, index=False, sheet_name="规则命中明细")
        pd.DataFrame(entries).to_excel(writer, index=False, sheet_name="最终发放")

    sample_ids = [
        "26448788",
        "27057228",
        "27535527",
        "27801891",
        "28030880",
        "28077470",
        "28112864",
        "28167026",
        "28438826",
        "28658041",
        "28658360",
        "28661515",
        "28662770",
    ]
    entry_ids = {row["student_id"] for row in entries}
    all_rule_ids = {str(row.get("student_id", "")) for row in audit_rows}
    summary = {
        "source": str(SOURCE),
        "source_size": SOURCE.stat().st_size,
        "source_rows": int(len(df)),
        "tw1_positive": int(sum(1 for row in entries if row["source"] == "台湾方案1")),
        "tw2": int(sum(1 for row in entries if row["source"] == "台湾方案2")),
        "intl": int(sum(1 for row in entries if row["source"] == "欧美澳港澳")),
        "total_issue_rows": len(entries),
        "coin_distribution": {
            str(k): int(v)
            for k, v in pd.Series([row["coins"] for row in entries]).value_counts().sort_index().items()
        },
        "missing_orders": int(sum(1 for row in entries if not row["order_no"] or row["order_no"].lower() == "nan")),
        "sample_in_issue": {sid: sid in entry_ids for sid in sample_ids},
        "sample_in_any_rule": {sid: sid in all_rule_ids for sid in sample_ids},
        "target": str(TARGET_XLS),
        "audit": str(AUDIT),
    }
    print(json.dumps(summary, ensure_ascii=True, indent=2))


if __name__ == "__main__":
    main()

